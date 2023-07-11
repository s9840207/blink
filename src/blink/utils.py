import pandas as pd


class OpenpyxlReader(pd.io.excel._openpyxl.OpenpyxlReader):
    def __init__(self, filepath_or_buffer):
        super().__init__(filepath_or_buffer)

    def load_workbook(self, filepath_or_buffer):
        from openpyxl import load_workbook

        return load_workbook(
            filepath_or_buffer, read_only=False, data_only=True, keep_links=False
        )


def read_excel(filepath, sheet_name=0):
    return OpenpyxlReader(filepath).parse(sheet_name=sheet_name)
